import asyncio
import os
import random
import threading

from flask import Flask
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    CallbackQueryHandler,
    filters,
)
from telegram.request import HTTPXRequest
from openpyxl import Workbook, load_workbook

# ======================
# WEB SERVER (ANTI-SLEEP)
# ======================
app_web = Flask(__name__)

@app_web.route('/')
def home():
    return "Bot is running"

def run_web():
    app_web.run(host='0.0.0.0', port=10000)

threading.Thread(target=run_web).start()

# ======================
# CONFIG
# ======================
TOKEN = os.getenv("BOT_TOKEN")   # 🔥 SECURE TOKEN
ADMIN_ID = 6654835556
TICKET_PRICE = 200
FILE_NAME = "lottery.xlsx"

ALL_TICKETS = [f"EMUN-{str(i).zfill(3)}" for i in range(1, 550)]
temp_user = {}

# ======================
# KEYBOARD
# ======================
def yes_no_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🟢 YES / እወ ይደሊ", callback_data="yes"),
            InlineKeyboardButton("🔴 NO", callback_data="no")
        ]
    ])

# ======================
# EXCEL
# ======================
def init_excel():
    try:
        load_workbook(FILE_NAME)
    except:
        wb = Workbook()
        ws = wb.active
        ws.append(["Phone", "Lottery Number", "Total", "Status"])
        wb.save(FILE_NAME)

def get_used_tickets():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    used = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == "approved":
            used.add(row[1])
    return used

def assign_tickets(count):
    used = get_used_tickets()
    available = list(set(ALL_TICKETS) - used)
    if len(available) < count:
        return None
    return random.sample(available, count)

def save_to_excel(phone, tickets):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    for t in tickets:
        ws.append([phone, t, TICKET_PRICE, "approved"])
    wb.save(FILE_NAME)

def remaining_count():
    return len(set(ALL_TICKETS) - get_used_tickets())

# ======================
# BOT HANDLERS
# ======================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🎉 Do you want a ticket?",
        reply_markup=yes_no_keyboard()
    )

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data

    # ADMIN
    if data.startswith("approve_") or data.startswith("reject_"):
        if user_id != ADMIN_ID:
            return await query.answer("❌ Not allowed", show_alert=True)

        target_id = int(data.split("_")[1])

        if target_id not in temp_user:
            return await query.edit_message_text("❗ Data missing")

        user = temp_user[target_id]

        if data.startswith("approve_"):
            save_to_excel(user["phone"], user["assigned_tickets"])

            await context.bot.send_message(
                target_id,
                "✅ Payment Confirmed!\n\n🎟 Tickets:\n" +
                "\n".join(user["assigned_tickets"])
            )

            await context.bot.send_message(
                target_id,
                "Want more tickets?",
                reply_markup=yes_no_keyboard()
            )

            await query.edit_message_text("✅ Approved")

        else:
            await context.bot.send_message(target_id, "❌ Payment Rejected")
            await query.edit_message_text("❌ Rejected")

        temp_user.pop(target_id, None)
        return

    # USER
    if data == "yes":
        temp_user[user_id] = {}
        await query.edit_message_text("📞 Enter phone number:")

    elif data == "no":
        await query.edit_message_text("👍 Okay!")

    elif data.startswith("paid_"):
        target_id = int(data.split("_")[1])
        user = temp_user.get(target_id)

        keyboard = [[
            InlineKeyboardButton("✅ Approve", callback_data=f"approve_{target_id}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"reject_{target_id}")
        ]]

        await context.bot.send_message(
            ADMIN_ID,
            f"📥 Payment Request\n\nPhone: {user['phone']}\n\n"
            + "\n".join(user["assigned_tickets"]) +
            f"\n\nTotal: {user['total']} birr",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

        await query.edit_message_text("⏳ Waiting for approval...")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text

    if user_id not in temp_user:
        return await update.message.reply_text("❗ Press /start first")

    if "phone" not in temp_user[user_id]:
        temp_user[user_id]["phone"] = text
        return await update.message.reply_text("How many tickets?")

    if "tickets" not in temp_user[user_id]:
        if not text.isdigit():
            return await update.message.reply_text("❗ Send number")

        count = int(text)
        tickets = assign_tickets(count)

        if not tickets:
            return await update.message.reply_text("❌ Not enough tickets")

        total = count * TICKET_PRICE

        temp_user[user_id]["tickets"] = count
        temp_user[user_id]["total"] = total
        temp_user[user_id]["assigned_tickets"] = tickets

        keyboard = [[InlineKeyboardButton("✅ I Paid", callback_data=f"paid_{user_id}")]]

        await update.message.reply_text(
            "\n".join(tickets) +
            f"\n\nTotal: {total} birr",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

# ======================
# MAIN LOOP (AUTO-RETRY)
# ======================
async def run_bot():
    init_excel()

    request = HTTPXRequest(connect_timeout=30, read_timeout=30)

    app = ApplicationBuilder().token(TOKEN).request(request).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(button))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    while True:
        try:
            print("Bot running...")
            await app.run_polling()
        except Exception as e:
            print("Error:", e)
            await asyncio.sleep(5)

if __name__ == "__main__":
    asyncio.run(run_bot())
